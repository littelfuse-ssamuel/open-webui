import asyncio
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference

from open_webui.routers import excel as excel_router


def _create_basic_workbook(file_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Category", "Value"])
    ws.append(["A", 10])
    ws.append(["B", 20])
    wb.save(file_path)
    wb.close()


def _create_chart_workbook(file_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Category", "Value"])
    ws.append(["A", 10])
    ws.append(["B", 20])

    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "E5")

    wb.save(file_path)
    wb.close()


def test_workbook_load_kwargs_uses_extension():
    assert excel_router._get_workbook_load_kwargs("macro.xlsm")["keep_vba"] is True
    assert excel_router._get_workbook_load_kwargs("report.xlsx")["keep_vba"] is False
    assert excel_router._get_workbook_load_kwargs("report.xlsx")["keep_links"] is True


def test_excel_update_preserves_chart_parts(tmp_path: Path):
    file_path = tmp_path / "chart.xlsx"
    _create_chart_workbook(file_path)

    changes = [excel_router.CellChange(row=2, col=2, value=15, isFormula=False)]
    applied = excel_router._apply_excel_changes_on_disk(
        file_path=file_path,
        filename="chart.xlsx",
        sheet_name="Sheet1",
        changes=changes,
        file_id="chart-file",
    )

    assert applied == 1

    with zipfile.ZipFile(file_path, "r") as zip_file:
        chart_parts = [
            name for name in zip_file.namelist() if name.startswith("xl/charts/chart")
        ]
        assert chart_parts

    wb = openpyxl.load_workbook(file_path, data_only=False)
    assert wb["Sheet1"]["B2"].value == 15
    wb.close()


def test_excel_update_lock_handles_concurrent_writes(tmp_path: Path):
    file_path = tmp_path / "concurrent.xlsx"
    _create_basic_workbook(file_path)

    change_a = [excel_router.CellChange(row=2, col=2, value=101, isFormula=False)]
    change_b = [excel_router.CellChange(row=3, col=2, value=202, isFormula=False)]

    async def _run_concurrent_writes():
        await asyncio.gather(
            excel_router._apply_excel_changes_with_lock(
                file_id="shared-lock",
                file_path=file_path,
                filename="concurrent.xlsx",
                sheet_name="Sheet1",
                changes=change_a,
            ),
            excel_router._apply_excel_changes_with_lock(
                file_id="shared-lock",
                file_path=file_path,
                filename="concurrent.xlsx",
                sheet_name="Sheet1",
                changes=change_b,
            ),
        )

    asyncio.run(_run_concurrent_writes())

    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb["Sheet1"]
    assert ws["B2"].value == 101
    assert ws["B3"].value == 202
    wb.close()


def test_excel_update_rolls_back_when_replace_fails(tmp_path: Path, monkeypatch):
    file_path = tmp_path / "rollback.xlsx"
    _create_basic_workbook(file_path)
    original_bytes = file_path.read_bytes()

    real_replace = excel_router.os.replace

    def failing_replace(src, dst):
        if Path(dst) == file_path:
            raise OSError("simulated replace failure")
        return real_replace(src, dst)

    monkeypatch.setattr(excel_router.os, "replace", failing_replace)

    with pytest.raises(OSError, match="simulated replace failure"):
        excel_router._apply_excel_changes_on_disk(
            file_path=file_path,
            filename="rollback.xlsx",
            sheet_name="Sheet1",
            changes=[excel_router.CellChange(row=2, col=2, value=50, isFormula=False)],
            file_id="rollback-file",
        )

    assert file_path.read_bytes() == original_bytes

    wb = openpyxl.load_workbook(file_path, data_only=False)
    assert wb["Sheet1"]["B2"].value == 10
    wb.close()


def test_formula_qc_detects_critical_invalid_tokens(tmp_path: Path):
    file_path = tmp_path / "qc_invalid.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=SUM(B1:B2)"
    ws["A2"] = "=#REF!+1"
    wb.save(file_path)
    wb.close()

    wb2 = openpyxl.load_workbook(file_path, data_only=False)
    issues, repairs_applied = excel_router._run_formula_qc_and_repairs(wb2)
    wb2.close()

    assert repairs_applied == 0
    assert any(i.issueType == "invalid_reference_token" and i.severity == "critical" for i in issues)


def test_formula_qc_auto_repairs_double_equals(tmp_path: Path):
    file_path = tmp_path / "qc_repair.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "==SUM(B1:B2)"
    wb.save(file_path)
    wb.close()

    wb2 = openpyxl.load_workbook(file_path, data_only=False)
    issues, repairs_applied = excel_router._run_formula_qc_and_repairs(wb2)
    assert wb2["Sheet1"]["A1"].value == "=SUM(B1:B2)"
    wb2.close()

    assert repairs_applied == 1
    assert any(i.issueType == "auto_repaired_formula_prefix" and i.severity == "warning" for i in issues)


def test_formula_qc_detects_missing_sheet_reference(tmp_path: Path):
    file_path = tmp_path / "qc_missing_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "='Missing Sheet'!B2+1"
    wb.save(file_path)
    wb.close()

    wb2 = openpyxl.load_workbook(file_path, data_only=False)
    issues, repairs_applied = excel_router._run_formula_qc_and_repairs(wb2)
    wb2.close()

    assert repairs_applied == 0
    assert any(i.issueType == "missing_sheet_reference" and i.severity == "critical" for i in issues)


def test_build_qc_report_blocks_on_critical():
    report = excel_router._build_qc_report(
        [
            excel_router.ExcelQcIssue(
                sheet="Sheet1",
                cell="A1",
                severity="critical",
                issueType="invalid_reference_token",
                message="bad",
            )
        ],
        operation="download",
    )

    assert report.blocked is True
    assert report.criticalUnresolved == 1
    assert report.blockReason


def test_resolve_llm_qc_model_id_prefers_request_over_valve_and_fallback():
    configured_models = {
        "request-model": {},
        "valve-model": {},
        "fallback-model": {},
    }

    model_id, source = excel_router._resolve_llm_qc_model_id(
        configured_models=configured_models,
        requested_model_id="request-model",
        valve_model_id="valve-model",
        fallback_model_id="fallback-model",
    )

    assert model_id == "request-model"
    assert source == "request"


def test_resolve_llm_qc_model_id_uses_valve_when_request_is_invalid():
    configured_models = {
        "valve-model": {},
        "fallback-model": {},
    }

    model_id, source = excel_router._resolve_llm_qc_model_id(
        configured_models=configured_models,
        requested_model_id="missing-request-model",
        valve_model_id="valve-model",
        fallback_model_id="fallback-model",
    )

    assert model_id == "valve-model"
    assert source == "valve"


def test_resolve_llm_qc_model_id_returns_none_when_no_candidates_are_configured():
    configured_models = {"configured": {}}

    model_id, source = excel_router._resolve_llm_qc_model_id(
        configured_models=configured_models,
        requested_model_id="missing-request-model",
        valve_model_id="missing-valve-model",
        fallback_model_id="missing-fallback-model",
    )

    assert model_id is None
    assert source is None


def test_preflight_blocks_formula_overwrite_in_strict_mode():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=SUM(B1:B2)"

    warnings = excel_router._collect_preflight_warnings(
        wb=wb,
        sheet_name="Sheet1",
        changes=[excel_router.CellChange(row=1, col=1, value=123)],
    )
    report = excel_router._build_preflight_report(
        warnings=warnings,
        strict_formula_mode=True,
        block_referenced_by_formula=False,
        allow_formula_overwrite=False,
    )
    wb.close()

    assert any(w.warning_type == "contains_formula" for w in warnings)
    assert report.safe_to_apply is False
    assert report.status == "blocked"


def test_preflight_allows_formula_overwrite_when_forced_per_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=SUM(B1:B2)"

    warnings = excel_router._collect_preflight_warnings(
        wb=wb,
        sheet_name="Sheet1",
        changes=[
            excel_router.CellChange(
                row=1,
                col=1,
                value=123,
                forceOverwriteFormula=True,
            )
        ],
    )
    report = excel_router._build_preflight_report(
        warnings=warnings,
        strict_formula_mode=True,
        block_referenced_by_formula=False,
        allow_formula_overwrite=False,
    )
    wb.close()

    assert report.safe_to_apply is True
    assert report.status == "ok"


def test_preflight_detects_cross_sheet_formula_dependencies():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = 10
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "=Sheet1!A1+1"

    warnings = excel_router._collect_preflight_warnings(
        wb=wb,
        sheet_name="Sheet1",
        changes=[excel_router.CellChange(row=1, col=1, value=11)],
    )
    report = excel_router._build_preflight_report(
        warnings=warnings,
        strict_formula_mode=False,
        block_referenced_by_formula=True,
        allow_formula_overwrite=False,
    )
    wb.close()

    dependency_warnings = [w for w in warnings if w.warning_type == "referenced_by_formula"]
    assert dependency_warnings
    assert "Sheet2!A1" in dependency_warnings[0].details["referenced_by"]
    assert report.safe_to_apply is False
    assert report.status == "blocked"


def test_qc_repairs_persist_when_saved_atomically(tmp_path: Path):
    file_path = tmp_path / "persist_repair.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "==SUM(B1:B2)"
    wb.save(file_path)
    wb.close()

    wb2 = openpyxl.load_workbook(
        file_path, **excel_router._get_workbook_load_kwargs("persist_repair.xlsx")
    )
    issues, repairs_applied = excel_router._run_formula_qc_and_repairs(wb2)
    assert any(i.issueType == "auto_repaired_formula_prefix" for i in issues)
    assert repairs_applied == 1

    excel_router._persist_workbook_atomically(
        wb=wb2,
        file_path=file_path,
        load_kwargs=excel_router._get_workbook_load_kwargs("persist_repair.xlsx"),
        file_id="persist-repair",
    )

    wb3 = openpyxl.load_workbook(file_path, data_only=False)
    assert wb3["Sheet1"]["A1"].value == "=SUM(B1:B2)"
    wb3.close()


def test_build_excel_metadata_update_uses_extension_content_type():
    qc_report = excel_router._build_qc_report([], operation="update")
    meta_xlsm = excel_router._build_excel_metadata_update(
        filename="macro.xlsm",
        sheet_names=["Sheet1"],
        active_sheet="Sheet1",
        changes_applied=2,
        qc_report=qc_report,
        operation="update",
        repairs_applied=1,
    )
    meta_xlsx = excel_router._build_excel_metadata_update(
        filename="book.xlsx",
        sheet_names=["Sheet1"],
        active_sheet="Sheet1",
        changes_applied=2,
        qc_report=qc_report,
        operation="update",
        repairs_applied=0,
    )

    assert meta_xlsm["content_type"] == "application/vnd.ms-excel.sheet.macroEnabled.12"
    assert (
        meta_xlsx["content_type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
