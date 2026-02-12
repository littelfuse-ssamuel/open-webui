from io import BytesIO

from openpyxl import load_workbook

from open_webui.excel.dfmea_formatter import build_dfmea_workbook_bytes


def _merged_ranges(ws):
    return {str(rng) for rng in ws.merged_cells.ranges}


def _rgb_suffix(color_value: str | None) -> str:
    if not color_value:
        return ""
    return color_value[-6:].upper()


def test_littelfuse_dfmea_layout_contains_expected_header_blocks():
    records = [
        {
            "Item / Function": "Power Stage",
            "Requirement": "Must operate under overcurrent condition",
            "Potential Failure Modes": "Fails open",
            "Potential Effect(s) of Failure": "Load interruption",
        },
        {
            "Item / Function": "Control Stage",
            "Requirement": "Must detect faults quickly",
            "Potential Failure Modes": "Fails to detect",
            "Potential Effect(s) of Failure": "Unsafe operation",
        },
    ]

    payload = build_dfmea_workbook_bytes(records, "littelfuse")
    wb = load_workbook(BytesIO(payload))
    ws = wb["DFMEA"]

    merged = _merged_ranges(ws)
    assert ws["B2"].value == "System"
    assert ws["F2"].value == "POTENTIAL FAILURE MODE AND EFFECTS ANALYSIS"
    assert "F2:M2" in merged
    assert "P12:T12" in merged
    assert ws.cell(row=12, column=2).value == "Item / Function"
    assert ws.row_dimensions[13].height == 70

    # Separator row between item groups should be tan.
    separator_row = 15
    assert _rgb_suffix(ws.cell(row=separator_row, column=2).fill.start_color.rgb) == "EBD7A9"


def test_carling_dfmea_layout_contains_expected_action_header_and_merges():
    records = [
        {
            "Item Functions": "Switching",
            "Potential Failure Modes": "Stuck closed",
            "Potential Effect(s) of Failure": "Continuous current flow",
        }
    ]

    payload = build_dfmea_workbook_bytes(records, "carling")
    wb = load_workbook(BytesIO(payload))
    ws = wb["DFMEA"]

    merged = _merged_ranges(ws)
    assert ws.cell(row=2, column=2).value == "Item Functions"
    assert ws["S2"].value == "After Actions Taken"
    assert "S2:W2" in merged
    assert ws.row_dimensions[2].height == 30
    assert ws.row_dimensions[3].height == 40


def test_unknown_template_defaults_to_littelfuse():
    payload = build_dfmea_workbook_bytes([], "unknown_template")
    wb = load_workbook(BytesIO(payload))
    ws = wb["DFMEA"]
    assert ws["B2"].value == "System"
