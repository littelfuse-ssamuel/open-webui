"""
DFMEA workbook formatter for Littelfuse and Carling templates.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# REFACTOR_TOUCHPOINT[OWUI_DELEGATION_PHASE2]: owner=open-webui; intent=port LF-specific DFMEA workbook formatting (littelfuse/carling) into open-webui artifact stack; fallback=emit simpler DFMEA workbook layout from tool layer.


def _set_border_sides(
    cell,
    *,
    left=None,
    right=None,
    top=None,
    bottom=None,
) -> None:
    current = cell.border
    cell.border = Border(
        left=left if left is not None else current.left,
        right=right if right is not None else current.right,
        top=top if top is not None else current.top,
        bottom=bottom if bottom is not None else current.bottom,
    )


def _copy_row_values(record: dict[str, Any], headers: list[str]) -> list[Any]:
    return [record.get(header, "") for header in headers]


def _write_rows(
    ws,
    records: list[dict[str, Any]],
    headers: list[str],
    *,
    start_row: int,
    start_col: int,
) -> None:
    for offset, record in enumerate(records):
        row_idx = start_row + offset
        values = _copy_row_values(record, headers)
        for col_offset, value in enumerate(values):
            ws.cell(row=row_idx, column=start_col + col_offset, value=value)


def _insert_littelfuse_separators(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    records_with_separators: list[dict[str, Any]] = []
    current_item = None

    for record in records:
        item_function = str(record.get("Item / Function", "") or "")
        if item_function and item_function != current_item:
            if current_item is not None:
                records_with_separators.append({"Item / Function": "SEPARATOR_ROW"})
            current_item = item_function
        records_with_separators.append(record)

    return records_with_separators


def _build_carling_template(ws, records: list[dict[str, Any]]) -> None:
    pre_action_headers = [
        "Item Functions",
        "Potential Failure Modes",
        "Max Sev",
        "Potential Effect(s) of Failure",
        "Sev",
        "Class",
        "Potential Cause / Mechanism of Failure",
        "Occ",
        "Current Design Controls (Prevention)",
        "Current Design Controls (Detection)",
        "Det",
        "RPN",
        "Severity x Occurence",
        "Severity x Detection",
        "Recommendations",
        "Responsibility",
        "Recommendation Status",
    ]
    post_action_headers = ["Post Sev", "Post Occ", "Post Det", "Post RPN", "% Reduction"]
    all_headers = pre_action_headers + post_action_headers

    normalized = [{header: row.get(header, "") for header in all_headers} for row in records]
    _write_rows(ws, normalized, all_headers, start_row=4, start_col=2)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007E3A", end_color="007E3A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    middle_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    light_green_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
    light_blue_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    thick_side = Side(style="thick")
    thick_top_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=thick_side,
        bottom=Side(style="thin"),
    )

    for col_idx, header_text in enumerate(pre_action_headers, start=2):
        cell1 = ws.cell(row=2, column=col_idx, value=header_text)
        cell2 = ws.cell(row=3, column=col_idx)
        cell1.border = thin_border
        cell2.border = thin_border
        ws.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
        cell1.font = header_font
        cell1.fill = header_fill
        cell1.alignment = header_align

    post_action_main_header_cell = ws.cell(row=2, column=19, value="After Actions Taken")
    ws.merge_cells("S2:W2")
    post_action_main_header_cell.font = header_font
    post_action_main_header_cell.fill = header_fill
    post_action_main_header_cell.alignment = header_align
    post_action_main_header_cell.border = thin_border

    post_action_sub_headers = ["Sev", "Occ", "Det", "RPN", "% Reduction"]
    for col_idx, header_text in enumerate(post_action_sub_headers, start=19):
        cell = ws.cell(row=3, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    widths = {
        "A": 3,
        "B": 15.71,
        "C": 16.71,
        "D": 6.71,
        "E": 13.71,
        "F": 6.71,
        "G": 7.71,
        "H": 16.71,
        "I": 6.71,
        "J": 15.71,
        "K": 15.71,
        "L": 6.71,
        "M": 6.71,
        "N": 11.71,
        "O": 11.71,
        "P": 17.71,
        "Q": 14.71,
        "R": 18.71,
        "S": 6.71,
        "T": 6.71,
        "U": 6.71,
        "V": 6.71,
        "W": 13.71,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 40

    max_data_row = max(4, ws.max_row)
    for row in ws.iter_rows(min_row=4, min_col=2, max_row=max_data_row, max_col=23):
        border_to_apply = thick_top_border if row[0].row == 4 else thin_border
        for cell in row:
            cell.border = border_to_apply
            cell.alignment = middle_left_align
            if cell.column_letter in ["B", "C", "D", "E", "F", "G"]:
                cell.alignment = center_align
            if cell.column_letter in ["D", "F", "I", "L", "S", "T", "U"]:
                cell.fill = light_green_fill
            elif cell.column_letter in ["M", "N", "O", "V"]:
                cell.fill = light_blue_fill

    def apply_merges(driving_col_idx: int, cols_to_merge: list[int]) -> None:
        start_row = 4
        for row_num in range(5, ws.max_row + 1):
            if ws.cell(row=row_num, column=driving_col_idx).value:
                end_row = row_num - 1
                if end_row > start_row:
                    for col_idx in cols_to_merge:
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=col_idx,
                            end_row=end_row,
                            end_column=col_idx,
                        )
                start_row = row_num
        if ws.max_row > start_row:
            for col_idx in cols_to_merge:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col_idx,
                    end_row=ws.max_row,
                    end_column=col_idx,
                )

    apply_merges(driving_col_idx=2, cols_to_merge=[2])
    apply_merges(driving_col_idx=3, cols_to_merge=[3, 4, 7])

    for row_num in range(5, ws.max_row + 1):
        if ws.cell(row=row_num, column=2).value:
            for col_idx in range(2, 24):
                cell = ws.cell(row=row_num - 1, column=col_idx)
                _set_border_sides(cell, bottom=thick_side)

    min_row, max_row = 2, ws.max_row
    min_col, max_col = 2, 23

    for col_idx in range(min_col, max_col + 1):
        _set_border_sides(ws.cell(min_row, col_idx), top=thick_side)
        _set_border_sides(ws.cell(max_row, col_idx), bottom=thick_side)

    for row_idx in range(min_row, max_row + 1):
        _set_border_sides(ws.cell(row_idx, min_col), left=thick_side)
        _set_border_sides(ws.cell(row_idx, max_col), right=thick_side)


def _build_littelfuse_template(ws, records: list[dict[str, Any]]) -> None:
    records_with_separators = _insert_littelfuse_separators(records)

    pre_action_headers = [
        "Item / Function",
        "Requirement",
        "Potential Failure Modes",
        "Potential Effect(s) of Failure",
        "Severity",
        "Classification",
        "Potential Cause(s) of Failure",
        "Controls Prevention",
        "Occurance",
        "Controls Detection",
        "Detection",
        "RPN",
        "Recommended Action(s)",
        "Responsibility & Target Completion Date",
    ]
    post_action_headers = [
        "Actions Taken / Completion Date",
        "Severity",
        "Occurance",
        "Detection",
        "RPN",
    ]
    all_headers = pre_action_headers + post_action_headers

    normalized = [{header: row.get(header, "") for header in all_headers} for row in records_with_separators]
    main_header_start_row = 12
    data_start_row = 14
    _write_rows(ws, normalized, all_headers, start_row=data_start_row, start_col=2)

    thin_border_side = Side(style="thin")
    thick_side = Side(style="thick")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )
    bottom_border = Border(bottom=thin_border_side)
    header_fill = PatternFill(start_color="007E3A", end_color="007E3A", fill_type="solid")
    tan_fill = PatternFill(start_color="ebd7a9", end_color="ebd7a9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    italic_bold_font = Font(italic=True, bold=True)
    right_align = Alignment(horizontal="right")
    vertical_align = Alignment(
        text_rotation=90, horizontal="center", vertical="bottom", wrap_text=False
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["B2"] = "System"
    ws["B4"] = "Subsystem"
    ws["B6"] = "Component"
    ws["B8"] = "Model Year(s)"
    ws["B10"] = "Core Team"
    ws["F2"] = "POTENTIAL FAILURE MODE AND EFFECTS ANALYSIS"
    ws["F2"].font = italic_bold_font
    ws["F2"].alignment = center_align
    ws["F4"] = "(DESIGN FMEA)"
    ws["F4"].font = bold_font
    ws["F4"].alignment = center_align
    ws["F6"] = "Design Responsibility"
    ws["F8"] = "Key Date"
    ws["O4"] = "FMEA Number:"
    ws["O4"].alignment = right_align
    ws["O6"] = "Prepared by:"
    ws["O6"].alignment = right_align
    ws["O8"] = "FMEA Date (Orig.):"
    ws["O8"].alignment = right_align

    merges_and_borders = {
        "F2:M2": None,
        "F4:M4": None,
        "F6:H6": None,
        "F8:H8": None,
        "O4:P4": None,
        "O6:P6": None,
        "O8:P8": None,
        "C2:D2": bottom_border,
        "C4:D4": bottom_border,
        "C6:D6": bottom_border,
        "C8:D8": bottom_border,
        "C10:M10": bottom_border,
        "I6:M6": bottom_border,
        "I8:M8": bottom_border,
        "Q4:T4": bottom_border,
        "Q6:T6": bottom_border,
        "Q8:T8": bottom_border,
    }
    for merge_range, border_style in merges_and_borders.items():
        ws.merge_cells(merge_range)
        if border_style:
            for row in ws[merge_range]:
                for cell in row:
                    cell.border = border_style
                    cell.alignment = center_align

    for row_num in [3, 5, 7, 9, 11]:
        ws.merge_cells(f"B{row_num}:T{row_num}")

    header_row_1 = main_header_start_row
    header_row_2 = main_header_start_row + 1

    for col_idx, text in enumerate(pre_action_headers, start=2):
        cell = ws.cell(row=header_row_1, column=col_idx, value=text)
        ws.merge_cells(
            start_row=header_row_1,
            start_column=col_idx,
            end_row=header_row_2,
            end_column=col_idx,
        )
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.cell(row=header_row_2, column=col_idx).border = thin_border

    ws.merge_cells(start_row=header_row_1, start_column=16, end_row=header_row_1, end_column=20)
    cell = ws.cell(row=header_row_1, column=16, value="Action Results")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    for col_idx in range(16, 21):
        ws.cell(row=header_row_1, column=col_idx).border = thin_border

    for col_idx, text in enumerate(post_action_headers, start=16):
        post_cell = ws.cell(row=header_row_2, column=col_idx, value=text)
        post_cell.font = header_font
        post_cell.fill = header_fill
        post_cell.alignment = center_align
        post_cell.border = thin_border

    vertical_text_cols = [6, 7, 10, 12, 17, 18, 19, 20]
    for col_idx in vertical_text_cols:
        row_idx = header_row_1 if col_idx <= 15 else header_row_2
        ws.cell(row=row_idx, column=col_idx).alignment = vertical_align

    widths = [
        3,
        15.71,
        24.71,
        16.71,
        14.71,
        4.71,
        4.71,
        15.71,
        15.71,
        4.71,
        14.71,
        4.71,
        5.71,
        14.71,
        14.71,
        15.71,
        4.71,
        4.71,
        4.71,
        5.71,
    ]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    heights = [15, 5, 15, 5, 15, 5, 15, 5, 15, 10, 15, 70]
    for idx, height in enumerate(heights, start=1):
        ws.row_dimensions[idx + 1].height = height

    num_cols = len(all_headers)
    for row_idx in range(data_start_row, ws.max_row + 1):
        first_cell = ws.cell(row=row_idx, column=2)
        if first_cell.value == "SEPARATOR_ROW":
            for col_idx in range(2, num_cols + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = tan_fill
                cell.border = thin_border
            ws.merge_cells(
                start_row=row_idx,
                start_column=2,
                end_row=row_idx,
                end_column=num_cols + 1,
            )
            first_cell.value = None
            ws.row_dimensions[row_idx].height = 10
        else:
            for col_idx in range(2, num_cols + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align

    data_row_indices: list[int] = []
    for row_idx in range(data_start_row, ws.max_row + 1):
        if ws.cell(row=row_idx, column=2).value is not None:
            data_row_indices.append(row_idx)

    def apply_merges(driving_col_idx: int, cols_to_merge: list[int], valid_rows: list[int]) -> None:
        if not valid_rows:
            return

        start_row = valid_rows[0]
        for idx in range(1, len(valid_rows)):
            row_num = valid_rows[idx]
            if ws.cell(row=row_num, column=driving_col_idx).value:
                end_row = valid_rows[idx - 1]
                if end_row > start_row:
                    for col_idx in cols_to_merge:
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=col_idx,
                            end_row=end_row,
                            end_column=col_idx,
                        )
                        ws.cell(row=start_row, column=col_idx).alignment = center_align
                start_row = row_num

        last_data_row = valid_rows[-1]
        if last_data_row > start_row:
            for col_idx in cols_to_merge:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col_idx,
                    end_row=last_data_row,
                    end_column=col_idx,
                )

    apply_merges(driving_col_idx=2, cols_to_merge=[2], valid_rows=data_row_indices)
    apply_merges(driving_col_idx=3, cols_to_merge=[3, 4], valid_rows=data_row_indices)

    max_col_letter = get_column_letter(num_cols + 1)

    for row in ws["B2:T11"]:
        for cell in row:
            if cell.row == 2:
                _set_border_sides(cell, top=thick_side)
            if cell.column == 2:
                _set_border_sides(cell, left=thick_side)
            if cell.column == 20:
                _set_border_sides(cell, right=thick_side)

    for row in ws[f"B{main_header_start_row}:{max_col_letter}{ws.max_row}"]:
        for cell in row:
            left = thick_side if cell.column == 2 else None
            right = thick_side if cell.column == num_cols + 1 else None
            top = thick_side if cell.row == header_row_1 else None
            bottom = thick_side if cell.row == 13 or cell.row == ws.max_row else None
            _set_border_sides(cell, left=left, right=right, top=top, bottom=bottom)


def build_dfmea_workbook_bytes(
    records: list[dict[str, Any]],
    template_name: str = "littelfuse",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "DFMEA"

    template_key = (template_name or "littelfuse").strip().lower()
    if template_key == "carling":
        _build_carling_template(ws, records)
    else:
        _build_littelfuse_template(ws, records)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
