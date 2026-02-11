"""
Single-process multi-pass orchestration for high-quality Excel generation.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any, Callable, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from .generation_spec import (
    ExcelGenerateRequest,
    GenerationScore,
    TemplateName,
    WorkbookChartPlan,
    WorkbookColumnSpec,
    WorkbookContentPlan,
    WorkbookDataSchema,
    WorkbookLayoutPlan,
    WorkbookQcPlan,
    WorkbookSheetSchema,
    WorkbookSpec,
    WorkbookSpecSummary,
    WorkbookStylePlan,
)


@dataclass
class GenerationExecutionResult:
    workbook: Workbook
    spec: WorkbookSpec
    score: GenerationScore
    refinement_iterations: int
    qc_issues: list[Any]
    repairs_applied: int
    qc_blocked: bool


class ExcelGenerationOrchestrator:
    """
    Implements staged generation:
    parse intent -> data plan -> layout plan -> content plan -> style plan -> build -> QC loop.
    """

    def execute(
        self,
        request: ExcelGenerateRequest,
        qc_evaluator: Callable[[Workbook], tuple[list[Any], int]],
    ) -> GenerationExecutionResult:
        spec = self.build_spec(request)
        workbook = self.build_workbook(spec, include_sample_data=request.include_sample_data)

        qc_issues, repairs_applied = qc_evaluator(workbook)
        score = self._compute_generation_score(
            workbook=workbook,
            spec=spec,
            qc_issues=qc_issues,
        )

        refinement_iterations = 0
        max_refinements = spec.qc_plan.max_refinement_iterations
        while refinement_iterations < max_refinements:
            if self._has_critical_issue(qc_issues):
                break
            if score.overall_score >= spec.qc_plan.minimum_visual_score:
                break

            refinement_iterations += 1
            self._apply_refinement_pass(workbook, spec, refinement_iterations)
            qc_issues, new_repairs = qc_evaluator(workbook)
            repairs_applied += int(new_repairs)
            score = self._compute_generation_score(
                workbook=workbook,
                spec=spec,
                qc_issues=qc_issues,
            )

        qc_blocked = bool(spec.qc_plan.block_on_critical and self._has_critical_issue(qc_issues))
        return GenerationExecutionResult(
            workbook=workbook,
            spec=spec,
            score=score,
            refinement_iterations=refinement_iterations,
            qc_issues=qc_issues,
            repairs_applied=repairs_applied,
            qc_blocked=qc_blocked,
        )

    def build_spec(self, request: ExcelGenerateRequest) -> WorkbookSpec:
        intent = self._parse_intent(request.prompt)
        data_schema = self._design_data_schema(
            template=request.template,
            max_rows=request.max_rows_per_sheet,
        )
        layout_plan = self._design_layout_plan(request.template, data_schema)
        content_plan = self._design_content_plan(
            prompt=request.prompt,
            template=request.template,
            data_schema=data_schema,
            layout_plan=layout_plan,
            include_charts=request.include_charts,
        )
        style_plan = self._design_style_plan(request.template)
        qc_plan = self._design_qc_plan(request)

        return WorkbookSpec(
            intent=intent,
            template=request.template,
            data_schema=data_schema,
            layout_plan=layout_plan,
            content_plan=content_plan,
            style_plan=style_plan,
            qc_plan=qc_plan,
        )

    def build_workbook(self, spec: WorkbookSpec, include_sample_data: bool = True) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_schema in spec.data_schema.sheets:
            ws = wb.create_sheet(title=sheet_schema.name)
            self._write_sheet_content(
                ws=ws,
                sheet_schema=sheet_schema,
                spec=spec,
                include_sample_data=include_sample_data,
            )

        self._create_charts(wb, spec.content_plan.chart_plans)
        return wb

    def build_spec_summary(
        self, spec: WorkbookSpec, refinement_iterations: int
    ) -> WorkbookSpecSummary:
        column_count = sum(len(sheet.columns) for sheet in spec.data_schema.sheets)
        return WorkbookSpecSummary(
            template=spec.template,
            sheet_count=len(spec.data_schema.sheets),
            column_count=column_count,
            has_charts=len(spec.content_plan.chart_plans) > 0,
            refinement_iterations=refinement_iterations,
        )

    def _parse_intent(self, prompt: str) -> str:
        lower_prompt = (prompt or "").lower()
        if "forecast" in lower_prompt or "budget" in lower_prompt:
            return "planning"
        if "operations" in lower_prompt or "manufacturing" in lower_prompt:
            return "operations_tracking"
        if "finance" in lower_prompt or "ledger" in lower_prompt:
            return "financial_reporting"
        if "dashboard" in lower_prompt:
            return "executive_dashboard"
        return "general_business_report"

    def _design_data_schema(self, template: TemplateName, max_rows: int) -> WorkbookDataSchema:
        row_count = max(3, min(max_rows, 200))
        if template == "finance":
            return WorkbookDataSchema(
                sheets=[
                    WorkbookSheetSchema(
                        name="Finance Summary",
                        row_count=6,
                        columns=[
                            WorkbookColumnSpec(name="Metric", type="text"),
                            WorkbookColumnSpec(name="Amount", type="currency"),
                            WorkbookColumnSpec(name="Variance", type="percentage"),
                        ],
                    ),
                    WorkbookSheetSchema(
                        name="Ledger",
                        row_count=row_count,
                        columns=[
                            WorkbookColumnSpec(name="Date", type="date"),
                            WorkbookColumnSpec(name="Account", type="text"),
                            WorkbookColumnSpec(name="Category", type="text"),
                            WorkbookColumnSpec(name="Amount", type="currency"),
                            WorkbookColumnSpec(
                                name="Running Balance",
                                type="formula",
                                formula="=IF(ROW()={data_start_row},D{row},E{prev_row}+D{row})",
                            ),
                        ],
                    ),
                ],
                relationships=["Ledger!E:E -> Finance Summary!B:B (rollup)"],
            )

        if template == "operations":
            return WorkbookDataSchema(
                sheets=[
                    WorkbookSheetSchema(
                        name="Operations Dashboard",
                        row_count=6,
                        columns=[
                            WorkbookColumnSpec(name="Metric", type="text"),
                            WorkbookColumnSpec(name="Current", type="number"),
                            WorkbookColumnSpec(name="Target", type="number"),
                            WorkbookColumnSpec(
                                name="Variance",
                                type="formula",
                                formula="=IFERROR((B{row}-C{row})/C{row},0)",
                            ),
                        ],
                    ),
                    WorkbookSheetSchema(
                        name="Operations Data",
                        row_count=row_count,
                        columns=[
                            WorkbookColumnSpec(name="Week", type="text"),
                            WorkbookColumnSpec(name="Units Produced", type="number"),
                            WorkbookColumnSpec(name="Defects", type="number"),
                            WorkbookColumnSpec(name="Downtime Hours", type="number"),
                            WorkbookColumnSpec(
                                name="Yield",
                                type="formula",
                                formula="=IFERROR((B{row}-C{row})/B{row},0)",
                            ),
                        ],
                    ),
                ],
                relationships=["Operations Data!E:E -> Operations Dashboard!B:B (average)"],
            )

        return WorkbookDataSchema(
            sheets=[
                WorkbookSheetSchema(
                    name="Executive Dashboard",
                    row_count=6,
                    columns=[
                        WorkbookColumnSpec(name="KPI", type="text"),
                        WorkbookColumnSpec(name="Current", type="number"),
                        WorkbookColumnSpec(name="Target", type="number"),
                        WorkbookColumnSpec(
                            name="Attainment",
                            type="formula",
                            formula="=IFERROR(B{row}/C{row},0)",
                        ),
                    ],
                ),
                WorkbookSheetSchema(
                    name="Sales Data",
                    row_count=row_count,
                    columns=[
                        WorkbookColumnSpec(name="Period", type="text"),
                        WorkbookColumnSpec(name="Revenue", type="currency"),
                        WorkbookColumnSpec(name="Cost", type="currency"),
                        WorkbookColumnSpec(
                            name="Margin",
                            type="formula",
                            formula="=IFERROR((B{row}-C{row})/B{row},0)",
                        ),
                    ],
                ),
            ],
            relationships=["Sales Data!D:D -> Executive Dashboard!D:D (summary)"],
        )

    def _design_layout_plan(
        self, template: TemplateName, data_schema: WorkbookDataSchema
    ) -> WorkbookLayoutPlan:
        freeze_panes = {}
        auto_filter_sheets = []
        column_widths = {}

        for sheet in data_schema.sheets:
            freeze_panes[sheet.name] = "A5"
            auto_filter_sheets.append(sheet.name)
            widths = {}
            for idx, column in enumerate(sheet.columns, start=1):
                if column.type in {"currency", "percentage", "number"}:
                    widths[get_column_letter(idx)] = 14
                elif column.type == "date":
                    widths[get_column_letter(idx)] = 13
                else:
                    widths[get_column_letter(idx)] = min(26, max(12, len(column.name) + 4))
            column_widths[sheet.name] = widths

        if template == "operations":
            for sheet_name in column_widths:
                column_widths[sheet_name]["A"] = 22

        return WorkbookLayoutPlan(
            header_row=4,
            data_start_row=5,
            freeze_panes=freeze_panes,
            auto_filter_sheets=auto_filter_sheets,
            column_widths=column_widths,
        )

    def _design_content_plan(
        self,
        prompt: str,
        template: TemplateName,
        data_schema: WorkbookDataSchema,
        layout_plan: WorkbookLayoutPlan,
        include_charts: bool,
    ) -> WorkbookContentPlan:
        title_map = {
            "executive_dashboard": "Executive Performance Dashboard",
            "finance": "Finance Performance Pack",
            "operations": "Operations Performance Dashboard",
        }
        subtitle = f"Auto-generated from prompt: {prompt[:120]}"

        summary_points = [
            "Use filters to drill into category-level insights.",
            "Review formula-driven metrics before publishing externally.",
            "Refresh source data monthly for trend reliability.",
        ]

        chart_plans: list[WorkbookChartPlan] = []
        if include_charts and len(data_schema.sheets) >= 2:
            source_sheet = data_schema.sheets[1]
            sheet_name = data_schema.sheets[0].name
            start = layout_plan.data_start_row
            end = layout_plan.data_start_row + max(2, min(source_sheet.row_count, 12)) - 1
            category_col = get_column_letter(1)
            value_col = get_column_letter(2)
            chart_type: str = "line" if template == "finance" else "bar"
            chart_plans.append(
                WorkbookChartPlan(
                    sheet=sheet_name,
                    chart_type=chart_type,  # type: ignore[arg-type]
                    title=f"{source_sheet.name} Trend",
                    data_sheet=source_sheet.name,
                    category_range=f"{category_col}{start}:{category_col}{end}",
                    value_range=f"{value_col}{start}:{value_col}{end}",
                    anchor="F4",
                )
            )

        return WorkbookContentPlan(
            workbook_title=title_map[template],
            workbook_subtitle=subtitle,
            summary_points=summary_points,
            chart_plans=chart_plans,
        )

    def _design_style_plan(self, template: TemplateName) -> WorkbookStylePlan:
        if template == "finance":
            return WorkbookStylePlan(
                theme_name="finance",
                primary_color="0B6E4F",
                secondary_color="DDEFEA",
                accent_color="1E9C70",
                zebra_fill_color="F4FBF8",
            )
        if template == "operations":
            return WorkbookStylePlan(
                theme_name="operations",
                primary_color="8B4A00",
                secondary_color="FCE8D4",
                accent_color="D17400",
                zebra_fill_color="FFF8EF",
            )
        return WorkbookStylePlan(
            theme_name="executive",
            primary_color="1F4E79",
            secondary_color="D9E2F3",
            accent_color="4472C4",
            zebra_fill_color="F7F9FC",
        )

    def _design_qc_plan(self, request: ExcelGenerateRequest) -> WorkbookQcPlan:
        return WorkbookQcPlan(
            block_on_critical=request.block_on_critical_qc,
            minimum_visual_score=request.minimum_visual_score,
            max_refinement_iterations=request.max_refinement_iterations,
        )

    def _write_sheet_content(
        self,
        ws,
        sheet_schema: WorkbookSheetSchema,
        spec: WorkbookSpec,
        include_sample_data: bool,
    ):
        layout = spec.layout_plan
        style = spec.style_plan

        title_font = Font(
            name=style.font_name,
            size=14,
            bold=True,
            color=style.primary_color,
        )
        subtitle_font = Font(
            name=style.font_name,
            size=10,
            italic=True,
            color=style.accent_color,
        )
        header_fill = PatternFill(
            fill_type="solid", start_color=style.primary_color, end_color=style.primary_color
        )
        header_font = Font(
            name=style.font_name, size=style.header_font_size, bold=True, color="FFFFFF"
        )
        header_border = Border(
            left=Side(style="thin", color="CFCFCF"),
            right=Side(style="thin", color="CFCFCF"),
            top=Side(style="thin", color="CFCFCF"),
            bottom=Side(style="thin", color="CFCFCF"),
        )

        ws["A1"] = spec.content_plan.workbook_title
        ws["A1"].font = title_font
        ws["A2"] = spec.content_plan.workbook_subtitle
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = Alignment(wrap_text=True)

        last_col = get_column_letter(max(1, len(sheet_schema.columns)))
        ws.merge_cells(f"A1:{last_col}1")

        if spec.content_plan.summary_points:
            ws["A3"] = spec.content_plan.summary_points[0]
            ws["A3"].font = Font(name=style.font_name, size=9, color=style.accent_color)

        for idx, column in enumerate(sheet_schema.columns, start=1):
            cell = ws.cell(row=layout.header_row, column=idx)
            cell.value = column.name
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if include_sample_data:
            for row_offset in range(sheet_schema.row_count):
                row = layout.data_start_row + row_offset
                prev_row = row - 1
                for col_idx, column in enumerate(sheet_schema.columns, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    value = self._sample_value(
                        column=column,
                        row_offset=row_offset,
                        row=row,
                        prev_row=prev_row,
                        data_start_row=layout.data_start_row,
                    )
                    cell.value = value
                    self._apply_number_format(cell, column)
                    if row_offset % 2 == 1:
                        cell.fill = PatternFill(
                            fill_type="solid",
                            start_color=style.zebra_fill_color,
                            end_color=style.zebra_fill_color,
                        )
                    cell.font = Font(name=style.font_name, size=style.body_font_size)
                    cell.border = header_border
                    if column.type in {"number", "currency", "percentage", "formula"}:
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")

        ws.freeze_panes = spec.layout_plan.freeze_panes.get(sheet_schema.name)
        if sheet_schema.name in layout.auto_filter_sheets:
            ws.auto_filter.ref = (
                f"A{layout.header_row}:{last_col}{layout.data_start_row + sheet_schema.row_count - 1}"
            )
        for col_letter, width in layout.column_widths.get(sheet_schema.name, {}).items():
            ws.column_dimensions[col_letter].width = width

    def _sample_value(
        self,
        column: WorkbookColumnSpec,
        row_offset: int,
        row: int,
        prev_row: int,
        data_start_row: int,
    ) -> Any:
        if column.sample_values:
            return column.sample_values[row_offset % len(column.sample_values)]
        if column.type == "date":
            return date.today() - timedelta(days=(30 - row_offset))
        if column.type == "text":
            if "period" in column.name.lower():
                months = [
                    "Jan",
                    "Feb",
                    "Mar",
                    "Apr",
                    "May",
                    "Jun",
                    "Jul",
                    "Aug",
                    "Sep",
                    "Oct",
                    "Nov",
                    "Dec",
                ]
                return months[row_offset % len(months)]
            if "week" in column.name.lower():
                return f"Week {row_offset + 1}"
            if "metric" in column.name.lower():
                metrics = [
                    "Revenue",
                    "Gross Margin",
                    "Operating Cost",
                    "Backlog",
                    "On-Time Delivery",
                    "Inventory Turns",
                ]
                return metrics[row_offset % len(metrics)]
            if "account" in column.name.lower():
                return f"Account {100 + row_offset}"
            if "category" in column.name.lower():
                categories = ["Core", "Strategic", "Support", "R&D"]
                return categories[row_offset % len(categories)]
            if "kpi" in column.name.lower():
                kpis = ["Revenue", "Margin", "NPS", "Retention", "Cash Conversion", "Backlog"]
                return kpis[row_offset % len(kpis)]
            return f"{column.name} {row_offset + 1}"
        if column.type == "number":
            return 100 + (row_offset * 13)
        if column.type == "currency":
            return 12000 + (row_offset * 1750)
        if column.type == "percentage":
            return min(0.98, 0.35 + (row_offset * 0.02))
        if column.type == "formula" and column.formula:
            return column.formula.format(
                row=row,
                prev_row=max(data_start_row, prev_row),
                data_start_row=data_start_row,
            )
        return None

    def _apply_number_format(self, cell, column: WorkbookColumnSpec):
        if column.type == "currency":
            cell.number_format = "$#,##0.00"
        elif column.type == "percentage":
            cell.number_format = "0.00%"
        elif column.type == "date":
            cell.number_format = "yyyy-mm-dd"

    def _create_charts(self, wb: Workbook, chart_plans: Sequence[WorkbookChartPlan]):
        chart_map = {"bar": BarChart, "line": LineChart, "pie": PieChart}
        for plan in chart_plans:
            if plan.sheet not in wb.sheetnames or plan.data_sheet not in wb.sheetnames:
                continue

            chart_ws = wb[plan.sheet]
            data_ws = wb[plan.data_sheet]
            min_col, min_row, max_col, max_row = range_boundaries(plan.value_range)
            cat_min_col, cat_min_row, cat_max_col, cat_max_row = range_boundaries(
                plan.category_range
            )
            chart = chart_map.get(plan.chart_type, BarChart)()
            if plan.chart_type == "bar":
                chart.type = "col"
            chart.title = plan.title
            values = Reference(
                data_ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
            )
            categories = Reference(
                data_ws,
                min_col=cat_min_col,
                min_row=cat_min_row,
                max_col=cat_max_col,
                max_row=cat_max_row,
            )
            chart.add_data(values, titles_from_data=False)
            chart.set_categories(categories)
            chart.height = 7
            chart.width = 11
            chart_ws.add_chart(chart, plan.anchor)

    def _apply_refinement_pass(self, wb: Workbook, spec: WorkbookSpec, iteration: int):
        for ws in wb.worksheets:
            header_row = spec.layout_plan.header_row
            max_col = ws.max_column
            if max_col <= 0:
                continue

            if iteration == 1:
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=header_row, column=col)
                    cell.font = Font(
                        name=spec.style_plan.font_name,
                        size=spec.style_plan.header_font_size + 1,
                        bold=True,
                        color="FFFFFF",
                    )
                    cell.fill = PatternFill(
                        fill_type="solid",
                        start_color=spec.style_plan.primary_color,
                        end_color=spec.style_plan.primary_color,
                    )
            else:
                for col in range(1, max_col + 1):
                    col_letter = get_column_letter(col)
                    current_width = ws.column_dimensions[col_letter].width or 12
                    ws.column_dimensions[col_letter].width = min(current_width + 2, 36)

                ws["A2"].font = Font(
                    name=spec.style_plan.font_name,
                    size=10,
                    italic=True,
                    color=spec.style_plan.accent_color,
                )

    def _compute_generation_score(
        self, workbook: Workbook, spec: WorkbookSpec, qc_issues: Iterable[Any]
    ) -> GenerationScore:
        issue_list = list(qc_issues)
        critical = [i for i in issue_list if self._issue_severity(i) == "critical"]
        warnings = [i for i in issue_list if self._issue_severity(i) == "warning"]

        structure_checks = []
        visual_checks = []

        structure_checks.append(len(workbook.sheetnames) >= len(spec.data_schema.sheets))
        for ws in workbook.worksheets:
            structure_checks.append(ws.max_column >= 1)
            structure_checks.append(ws.max_row >= spec.layout_plan.data_start_row)
            visual_checks.append(ws.freeze_panes is not None)
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=spec.layout_plan.header_row, column=col)
                visual_checks.append(header_cell.fill is not None and header_cell.fill.fill_type == "solid")
                visual_checks.append(bool(header_cell.font and header_cell.font.bold))

        if spec.content_plan.chart_plans:
            has_charts = any(len(getattr(ws, "_charts", []) or []) > 0 for ws in workbook.worksheets)
            visual_checks.append(has_charts)

        structure_score = self._ratio(structure_checks)
        visual_score = self._ratio(visual_checks)
        formula_score = max(0.0, 1.0 - (len(critical) * 0.35) - (len(warnings) * 0.05))
        overall_score = (
            (visual_score * 0.45)
            + (structure_score * 0.35)
            + (formula_score * 0.20)
        )
        overall_score = max(0.0, min(1.0, overall_score))

        return GenerationScore(
            visual_score=round(visual_score, 4),
            structure_score=round(structure_score, 4),
            formula_score=round(formula_score, 4),
            overall_score=round(overall_score, 4),
        )

    def _issue_severity(self, issue: Any) -> str:
        if isinstance(issue, dict):
            return str(issue.get("severity", "")).lower()
        return str(getattr(issue, "severity", "")).lower()

    def _has_critical_issue(self, issues: Iterable[Any]) -> bool:
        for issue in issues:
            if self._issue_severity(issue) == "critical":
                return True
        return False

    def _ratio(self, checks: Sequence[bool]) -> float:
        if not checks:
            return 0.0
        passed = sum(1 for check in checks if check)
        return passed / len(checks)

