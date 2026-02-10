"""
Excel file editing API endpoints
"""

import asyncio
import logging
import os
import re
import shutil
import tempfile
import zipfile
from contextlib import suppress
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional, Literal
from fastapi import APIRouter, Depends, HTTPException, status
from pydantic import BaseModel
import openpyxl
from openpyxl.utils import get_column_letter

from open_webui.constants import ERROR_MESSAGES
from open_webui.models.files import Files
from open_webui.storage.provider import Storage
from open_webui.utils.auth import get_verified_user

log = logging.getLogger(__name__)

router = APIRouter()

_EXCEL_FILE_LOCKS: dict[str, asyncio.Lock] = {}
_EXCEL_FILE_LOCKS_GUARD = asyncio.Lock()


async def _get_excel_file_lock(file_id: str) -> asyncio.Lock:
    async with _EXCEL_FILE_LOCKS_GUARD:
        lock = _EXCEL_FILE_LOCKS.get(file_id)
        if lock is None:
            lock = asyncio.Lock()
            _EXCEL_FILE_LOCKS[file_id] = lock
        return lock


def _get_workbook_load_kwargs(filename: str) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower()
    return {
        "keep_vba": suffix == ".xlsm",
        "data_only": False,
        "keep_links": True,
    }


def _remove_calc_chain_if_present(file_path: Path):
    if not zipfile.is_zipfile(file_path):
        return

    fd, tmp_name = tempfile.mkstemp(
        prefix=f"{file_path.stem}_calc_", suffix=file_path.suffix, dir=file_path.parent
    )
    os.close(fd)
    tmp_path = Path(tmp_name)

    should_replace = False
    try:
        with zipfile.ZipFile(file_path, "r") as zin:
            members = zin.namelist()
            if "xl/calcChain.xml" not in members:
                return

            with zipfile.ZipFile(tmp_path, "w") as zout:
                for item in zin.infolist():
                    if item.filename != "xl/calcChain.xml":
                        zout.writestr(item, zin.read(item.filename))
            should_replace = True

        if should_replace:
            os.replace(tmp_path, file_path)
    finally:
        with suppress(Exception):
            if tmp_path.exists():
                tmp_path.unlink()


def _coerce_cell_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value

    stripped = value.strip()
    if stripped == "":
        return value

    try:
        if "." in stripped:
            return float(stripped)
        return int(stripped)
    except ValueError:
        return value


def _apply_worksheet_changes(ws, changes: List["CellChange"]) -> int:
    changes_applied = 0

    for change in changes:
        try:
            if change.row < 1 or change.col < 1:
                log.warning(
                    f"Invalid cell coordinates: row={change.row}, col={change.col}"
                )
                continue

            if change.value is None:
                log.debug(
                    f"Skipping cell at row={change.row}, col={change.col} - value is None"
                )
                continue

            cell = ws.cell(row=change.row, column=change.col)

            if (
                change.isFormula
                and isinstance(change.value, str)
                and change.value.startswith("=")
            ):
                cell.value = change.value
                log.debug(
                    f"Updated cell {get_column_letter(change.col)}{change.row} with formula: {change.value}"
                )
            else:
                value = _coerce_cell_value(change.value)
                cell.value = value
                log.debug(
                    f"Updated cell {get_column_letter(change.col)}{change.row} = {value}"
                )

            changes_applied += 1

        except Exception as e:
            log.error(f"Error updating cell at row={change.row}, col={change.col}: {e}")
            continue

    return changes_applied


def _persist_workbook_atomically(
    wb, file_path: Path, load_kwargs: dict[str, Any], file_id: str
):
    fd, tmp_name = tempfile.mkstemp(
        prefix=f"{file_path.stem}_update_", suffix=file_path.suffix, dir=file_path.parent
    )
    os.close(fd)
    tmp_path = Path(tmp_name)
    backup_path = file_path.with_suffix(f"{file_path.suffix}.bak")

    try:
        try:
            try:
                wb.properties.calcId = None
            except Exception:
                pass
            wb.save(tmp_path)
        finally:
            with suppress(Exception):
                wb.close()

        _remove_calc_chain_if_present(tmp_path)

        verify_wb = openpyxl.load_workbook(tmp_path, **load_kwargs)
        verify_wb.close()

        shutil.copy2(file_path, backup_path)
        try:
            os.replace(tmp_path, file_path)
        except Exception:
            with suppress(Exception):
                shutil.copy2(backup_path, file_path)
            raise

        log.info(f"Successfully wrote updated workbook for file {file_id}")
    finally:
        with suppress(Exception):
            if tmp_path.exists():
                tmp_path.unlink()
        with suppress(Exception):
            if backup_path.exists():
                backup_path.unlink()


def _apply_excel_changes_on_disk(
    file_path: Path,
    filename: str,
    sheet_name: str,
    changes: List["CellChange"],
    file_id: str,
) -> int:
    load_kwargs = _get_workbook_load_kwargs(filename)
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, **load_kwargs)
    except Exception as e:
        log.error(f"Error loading workbook {file_id}: {e}")
        raise ValueError(f"Invalid Excel file: {str(e)}")

    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

        ws = wb[sheet_name]
        changes_applied = _apply_worksheet_changes(ws, changes)
        _persist_workbook_atomically(wb, file_path, load_kwargs, file_id)
        wb = None
        return changes_applied
    finally:
        if wb is not None:
            with suppress(Exception):
                wb.close()


async def _apply_excel_changes_with_lock(
    file_id: str,
    file_path: Path,
    filename: str,
    sheet_name: str,
    changes: List["CellChange"],
) -> int:
    lock = await _get_excel_file_lock(file_id)
    async with lock:
        return _apply_excel_changes_on_disk(
            file_path=file_path,
            filename=filename,
            sheet_name=sheet_name,
            changes=changes,
            file_id=file_id,
        )


class CellChange(BaseModel):
    """Represents a single cell change"""
    row: int  # 1-based row number
    col: int  # 1-based column number
    value: Optional[str | int | float | bool] = None
    isFormula: Optional[bool] = False  # Whether the value is a formula (starts with =)
    forceOverwriteFormula: Optional[bool] = False


class ExcelUpdateRequest(BaseModel):
    """Request to update cells in an Excel file"""
    fileId: str
    sheet: str
    changes: List[CellChange]
    strictFormulaMode: Optional[bool] = True
    blockReferencedByFormula: Optional[bool] = True
    allowFormulaOverwrite: Optional[bool] = False


class ExcelUpdateResponse(BaseModel):
    """Response from Excel update operation"""
    status: str
    message: Optional[str] = None
    preflightReport: Optional["ExcelValidationResponse"] = None
    qcReport: Optional["ExcelQcReport"] = None
    repairsApplied: Optional[int] = 0
    metadataUpdated: Optional[bool] = None


class ExcelQcIssue(BaseModel):
    sheet: str
    cell: str
    severity: Literal["critical", "warning"]
    issueType: str
    message: str
    originalFormula: Optional[str] = None
    repairedFormula: Optional[str] = None


class ExcelQcReport(BaseModel):
    blocked: bool
    blockReason: str
    criticalUnresolved: int
    issues: List[ExcelQcIssue]
    recommendedActions: List[str]


class ExcelDownloadReadyRequest(BaseModel):
    fileId: str
    strictMode: Optional[bool] = True
    allowLlmRepair: Optional[bool] = False
    llmModelId: Optional[str] = None
    valveLlmModelId: Optional[str] = None
    fallbackModelId: Optional[str] = None


class ExcelDownloadReadyResponse(BaseModel):
    status: str
    downloadUrl: Optional[str] = None
    qcReport: Optional[ExcelQcReport] = None
    selectedLlmModelId: Optional[str] = None
    selectedLlmModelSource: Optional[str] = None
    repairsApplied: Optional[int] = 0
    metadataUpdated: Optional[bool] = None


class ExcelMetadataResponse(BaseModel):
    """Response with Excel file metadata"""
    fileId: str
    sheetNames: List[str]
    activeSheet: Optional[str] = None


def _build_qc_report(issues: list[ExcelQcIssue], operation: str) -> ExcelQcReport:
    critical = [i for i in issues if i.severity == "critical"]
    blocked = len(critical) > 0
    reason = (
        "QC blocked: unresolved formula integrity issues"
        if blocked
        else "No blocking quality issues detected"
    )
    return ExcelQcReport(
        blocked=blocked,
        blockReason=reason,
        criticalUnresolved=len(critical),
        issues=issues[:50],
        recommendedActions=(
            [
                "Review impacted formulas and referenced ranges before retrying.",
                "Fix invalid references (#REF!, #NAME?, #VALUE!) in the listed cells.",
                f"Retry the {operation} after correcting the workbook formulas.",
            ]
            if blocked
            else []
        ),
    )


def _excel_content_type(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsm":
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_excel_metadata_update(
    filename: str,
    sheet_names: list[str],
    active_sheet: Optional[str],
    changes_applied: int,
    qc_report: ExcelQcReport,
    operation: str,
    repairs_applied: int,
) -> dict[str, Any]:
    return {
        "sheetNames": sheet_names,
        "activeSheet": active_sheet,
        "content_type": _excel_content_type(filename),
        "excel_last_operation": operation,
        "excel_last_edited_at": datetime.now().isoformat(),
        "excel_last_changes_applied": changes_applied,
        "excel_last_repairs_applied": repairs_applied,
        "excel_qc": {
            "blocked": qc_report.blocked,
            "criticalUnresolved": qc_report.criticalUnresolved,
            "issueCount": len(qc_report.issues),
            "blockReason": qc_report.blockReason,
        },
    }


def _update_excel_file_metadata(
    file_id: str,
    filename: str,
    sheet_names: list[str],
    active_sheet: Optional[str],
    changes_applied: int,
    qc_report: ExcelQcReport,
    operation: str,
    repairs_applied: int,
) -> bool:
    update_payload = _build_excel_metadata_update(
        filename=filename,
        sheet_names=sheet_names,
        active_sheet=active_sheet,
        changes_applied=changes_applied,
        qc_report=qc_report,
        operation=operation,
        repairs_applied=repairs_applied,
    )
    return Files.update_file_metadata_by_id(file_id, update_payload) is not None


def _run_formula_qc_and_repairs(wb) -> tuple[list[ExcelQcIssue], int]:
    issues: list[ExcelQcIssue] = []
    invalid_token_re = re.compile(r"#REF!|#NAME\?|#VALUE!", re.IGNORECASE)
    repairs_applied = 0
    sheets_by_lower = {sheet.lower(): sheet for sheet in wb.sheetnames}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue

                formula = value
                repaired_formula = formula
                repaired = False

                if formula.startswith("=="):
                    repaired_formula = "=" + formula.lstrip("=")
                    repaired = True

                if repaired_formula.count("(") != repaired_formula.count(")"):
                    issues.append(
                        ExcelQcIssue(
                            sheet=sheet_name,
                            cell=f"{get_column_letter(cell.column)}{cell.row}",
                            severity="critical",
                            issueType="unbalanced_parentheses",
                            message="Formula has unbalanced parentheses.",
                            originalFormula=formula,
                            repairedFormula=repaired_formula if repaired else None,
                        )
                    )

                if invalid_token_re.search(repaired_formula):
                    issues.append(
                        ExcelQcIssue(
                            sheet=sheet_name,
                            cell=f"{get_column_letter(cell.column)}{cell.row}",
                            severity="critical",
                            issueType="invalid_reference_token",
                            message="Formula contains an invalid Excel error token.",
                            originalFormula=formula,
                            repairedFormula=repaired_formula if repaired else None,
                        )
                    )

                refs = _parse_cell_references_from_formula(repaired_formula)
                missing_sheets = set()
                for ref in refs:
                    if "!" not in ref:
                        continue
                    referenced_sheet, _ = ref.split("!", 1)
                    normalized_sheet = _normalize_sheet_name(referenced_sheet)
                    if normalized_sheet.lower() not in sheets_by_lower:
                        missing_sheets.add(normalized_sheet)

                for missing_sheet in sorted(missing_sheets):
                    issues.append(
                        ExcelQcIssue(
                            sheet=sheet_name,
                            cell=f"{get_column_letter(cell.column)}{cell.row}",
                            severity="critical",
                            issueType="missing_sheet_reference",
                            message=(
                                "Formula references a sheet that does not exist: "
                                f"'{missing_sheet}'."
                            ),
                            originalFormula=formula,
                            repairedFormula=repaired_formula if repaired else None,
                        )
                    )

                if repaired and repaired_formula != formula:
                    cell.value = repaired_formula
                    repairs_applied += 1
                    issues.append(
                        ExcelQcIssue(
                            sheet=sheet_name,
                            cell=f"{get_column_letter(cell.column)}{cell.row}",
                            severity="warning",
                            issueType="auto_repaired_formula_prefix",
                            message="Auto-repaired malformed formula prefix.",
                            originalFormula=formula,
                            repairedFormula=repaired_formula,
                        )
                    )

    return issues, repairs_applied


def _resolve_llm_qc_model_id(
    configured_models: dict[str, Any],
    requested_model_id: Optional[str] = None,
    valve_model_id: Optional[str] = None,
    fallback_model_id: Optional[str] = None,
) -> tuple[Optional[str], Optional[str]]:
    candidates = [
        ("request", requested_model_id),
        ("valve", valve_model_id),
        ("fallback", fallback_model_id),
    ]

    for source, model_id in candidates:
        if model_id and model_id in configured_models:
            return model_id, source

    return None, None


@router.post("/update")
async def update_excel_file(
    request: ExcelUpdateRequest,
    user=Depends(get_verified_user),
) -> ExcelUpdateResponse:
    """
    Update cells in an Excel workbook.

    Args:
        request: The update request with file ID, sheet name, and cell changes
        user: The authenticated user

    Returns:
        ExcelUpdateResponse with status

    Raises:
        HTTPException: If file not found, access denied, or update fails
    """
    try:
        # Get the file from database
        file = Files.get_file_by_id(request.fileId)

        if not file:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ERROR_MESSAGES.NOT_FOUND,
            )

        # Check user permissions
        if file.user_id != user.id and user.role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
            )

        # Get the file path from storage
        file_path = Storage.get_file(file.path)
        file_path = Path(file_path)

        if not file_path.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found in storage",
            )

        preflight_report: Optional[ExcelValidationResponse] = None
        qc_report: Optional[ExcelQcReport] = None
        repairs_applied = 0
        metadata_updated: Optional[bool] = None

        try:
            lock = await _get_excel_file_lock(request.fileId)
            async with lock:
                wb = None
                try:
                    load_kwargs = _get_workbook_load_kwargs(file.filename)
                    wb = openpyxl.load_workbook(file_path, **load_kwargs)

                    if request.sheet not in wb.sheetnames:
                        raise ValueError(f"Sheet '{request.sheet}' not found in workbook.")

                    preflight_warnings = _collect_preflight_warnings(
                        wb=wb,
                        sheet_name=request.sheet,
                        changes=request.changes,
                    )
                    preflight_report = _build_preflight_report(
                        warnings=preflight_warnings,
                        strict_formula_mode=bool(request.strictFormulaMode),
                        block_referenced_by_formula=bool(
                            request.blockReferencedByFormula
                        ),
                        allow_formula_overwrite=bool(request.allowFormulaOverwrite),
                    )

                    if not preflight_report.safe_to_apply:
                        return ExcelUpdateResponse(
                            status="blocked",
                            message="Preflight blocked: formula-safety checks failed",
                            preflightReport=preflight_report,
                        )

                    ws = wb[request.sheet]
                    changes_applied = _apply_worksheet_changes(ws, request.changes)

                    qc_issues, repairs_applied = _run_formula_qc_and_repairs(wb)
                    qc_report = _build_qc_report(qc_issues, operation="save")

                    if qc_report.blocked:
                        return ExcelUpdateResponse(
                            status="blocked",
                            message="QC blocked: unresolved formula integrity issues",
                            preflightReport=preflight_report,
                            qcReport=qc_report,
                            repairsApplied=repairs_applied,
                            metadataUpdated=False,
                        )

                    sheet_names = wb.sheetnames
                    active_sheet = wb.active.title if wb.active else None
                    _persist_workbook_atomically(
                        wb=wb,
                        file_path=file_path,
                        load_kwargs=load_kwargs,
                        file_id=request.fileId,
                    )
                    wb = None

                    metadata_updated = _update_excel_file_metadata(
                        file_id=request.fileId,
                        filename=file.filename,
                        sheet_names=sheet_names,
                        active_sheet=active_sheet,
                        changes_applied=changes_applied,
                        qc_report=qc_report,
                        operation="update",
                        repairs_applied=repairs_applied,
                    )
                    if not metadata_updated:
                        log.warning(
                            "Excel metadata update failed for file %s after update",
                            request.fileId,
                        )
                finally:
                    if wb is not None:
                        with suppress(Exception):
                            wb.close()
        except ValueError as e:
            message = str(e)
            if message.startswith("Sheet '"):
                available_sheets: list[str] = []
                try:
                    wb_meta = openpyxl.load_workbook(
                        file_path, read_only=True, data_only=True, keep_links=True
                    )
                    available_sheets = wb_meta.sheetnames
                    wb_meta.close()
                except Exception:
                    available_sheets = []

                detail = (
                    f"{message} Available sheets: {', '.join(available_sheets)}"
                    if available_sheets
                    else message
                )
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=detail,
                )
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=message,
            )
        except Exception as e:
            log.error(f"Error saving workbook {request.fileId}: {e}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to save changes: {str(e)}",
            )

        return ExcelUpdateResponse(
            status="ok",
            message=f"Successfully updated {changes_applied} cells",
            preflightReport=preflight_report,
            qcReport=qc_report,
            repairsApplied=repairs_applied,
            metadataUpdated=metadata_updated,
        )

    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Unexpected error updating Excel file: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}",
        )




@router.post("/download-ready")
async def excel_download_ready(
    request: ExcelDownloadReadyRequest,
    user=Depends(get_verified_user),
) -> ExcelDownloadReadyResponse:
    """Run final QC gate before download and return allow/block state."""
    try:
        file = Files.get_file_by_id(request.fileId)

        if not file:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ERROR_MESSAGES.NOT_FOUND,
            )

        if file.user_id != user.id and user.role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
            )

        file_path = Path(Storage.get_file(file.path))
        if not file_path.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found in storage",
            )

        repairs_applied = 0
        metadata_updated: Optional[bool] = None
        qc_report: Optional[ExcelQcReport] = None
        lock = await _get_excel_file_lock(request.fileId)
        try:
            async with lock:
                wb = None
                try:
                    load_kwargs = _get_workbook_load_kwargs(file.filename)
                    wb = openpyxl.load_workbook(file_path, **load_kwargs)
                    issues, repairs_applied = _run_formula_qc_and_repairs(wb)
                    qc_report = _build_qc_report(issues, operation="download")

                    if qc_report.blocked:
                        metadata_updated = _update_excel_file_metadata(
                            file_id=request.fileId,
                            filename=file.filename,
                            sheet_names=wb.sheetnames,
                            active_sheet=wb.active.title if wb.active else None,
                            changes_applied=0,
                            qc_report=qc_report,
                            operation="download_ready_blocked",
                            repairs_applied=repairs_applied,
                        )
                        wb.close()
                        wb = None
                    elif repairs_applied > 0:
                        sheet_names = wb.sheetnames
                        active_sheet = wb.active.title if wb.active else None
                        _persist_workbook_atomically(
                            wb=wb,
                            file_path=file_path,
                            load_kwargs=load_kwargs,
                            file_id=request.fileId,
                        )
                        wb = None
                        metadata_updated = _update_excel_file_metadata(
                            file_id=request.fileId,
                            filename=file.filename,
                            sheet_names=sheet_names,
                            active_sheet=active_sheet,
                            changes_applied=0,
                            qc_report=qc_report,
                            operation="download_ready_repair",
                            repairs_applied=repairs_applied,
                        )
                    else:
                        metadata_updated = _update_excel_file_metadata(
                            file_id=request.fileId,
                            filename=file.filename,
                            sheet_names=wb.sheetnames,
                            active_sheet=wb.active.title if wb.active else None,
                            changes_applied=0,
                            qc_report=qc_report,
                            operation="download_ready_check",
                            repairs_applied=0,
                        )
                        wb.close()
                        wb = None
                finally:
                    if wb is not None:
                        with suppress(Exception):
                            wb.close()
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid Excel file: {str(e)}",
            )

        selected_llm_model_id = None
        selected_llm_model_source = None
        if request.allowLlmRepair:
            configured_models = request.app.state.MODELS or {}
            if not configured_models:
                try:
                    from open_webui.main import get_all_models

                    await get_all_models(request, user=user)
                    configured_models = request.app.state.MODELS or {}
                except Exception:
                    configured_models = request.app.state.MODELS or {}

            selected_llm_model_id, selected_llm_model_source = _resolve_llm_qc_model_id(
                configured_models=configured_models,
                requested_model_id=request.llmModelId,
                valve_model_id=request.valveLlmModelId,
                fallback_model_id=request.fallbackModelId,
            )

        if qc_report.blocked:
            log.info(
                "excel_qc_blocked fileId=%s operation=download blocked=true critical=%s selected_model=%s source=%s",
                request.fileId,
                qc_report.criticalUnresolved,
                selected_llm_model_id,
                selected_llm_model_source,
            )
            return ExcelDownloadReadyResponse(
                status="blocked",
                qcReport=qc_report,
                selectedLlmModelId=selected_llm_model_id,
                selectedLlmModelSource=selected_llm_model_source,
                repairsApplied=repairs_applied,
                metadataUpdated=metadata_updated,
            )

        return ExcelDownloadReadyResponse(
            status="ok",
            downloadUrl=f"/api/v1/files/{request.fileId}/content",
            qcReport=qc_report,
            selectedLlmModelId=selected_llm_model_id,
            selectedLlmModelSource=selected_llm_model_source,
            repairsApplied=repairs_applied,
            metadataUpdated=metadata_updated,
        )

    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Unexpected error in download-ready for {request.fileId}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}",
        )


@router.get("/{file_id}/metadata")
async def get_excel_metadata(
    file_id: str,
    user=Depends(get_verified_user),
) -> ExcelMetadataResponse:
    """
    Get metadata about an Excel file (sheet names, etc.).

    Args:
        file_id: The ID of the Excel file
        user: The authenticated user

    Returns:
        ExcelMetadataResponse with sheet names and active sheet

    Raises:
        HTTPException: If file not found or access denied
    """
    try:
        # Get the file from database
        file = Files.get_file_by_id(file_id)

        if not file:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ERROR_MESSAGES.NOT_FOUND,
            )

        # Check user permissions
        if file.user_id != user.id and user.role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
            )

        # Get the file path from storage
        file_path = Storage.get_file(file.path)
        file_path = Path(file_path)

        if not file_path.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found in storage",
            )

        # Load the workbook (data_only=True for faster loading)
        try:
            metadata_load_kwargs = {
                **_get_workbook_load_kwargs(file.filename),
                "data_only": True,
                "read_only": True,
            }
            wb = openpyxl.load_workbook(file_path, **metadata_load_kwargs)
            sheet_names = wb.sheetnames
            active_sheet = wb.active.title if wb.active else None
            wb.close()

            return ExcelMetadataResponse(
                fileId=file_id,
                sheetNames=sheet_names,
                activeSheet=active_sheet,
            )

        except Exception as e:
            log.error(f"Error reading workbook {file_id}: {e}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid Excel file: {str(e)}",
            )

    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Unexpected error getting Excel metadata: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}",
        )

class CellReference(BaseModel):
    """A cell reference with sheet, row, col"""
    sheet: str
    row: int
    col: int
    address: str  # e.g., "A1"


class CellWarning(BaseModel):
    """Warning about a cell that may be affected by changes"""
    cell: CellReference
    warning_type: str  # "contains_formula", "referenced_by_formula", "in_named_range"
    message: str
    details: Optional[dict] = None


class ExcelValidationRequest(BaseModel):
    """Request to validate proposed changes before applying"""
    fileId: str
    sheet: str
    changes: List[CellChange]
    strictFormulaMode: Optional[bool] = True
    blockReferencedByFormula: Optional[bool] = True
    allowFormulaOverwrite: Optional[bool] = False


class ExcelValidationResponse(BaseModel):
    """Response with validation warnings"""
    status: str
    warnings: List[CellWarning]
    safe_to_apply: bool
    message: Optional[str] = None


def _parse_cell_references_from_formula(formula: str) -> List[str]:
    """
    Extract cell references from a formula string.
    Returns list of cell addresses like ['A1', 'B2:B10', 'Sheet2!C3']
    """
    import re
    # Match cell references: A1, $A$1, A1:B10, Sheet1!A1, 'Sheet Name'!A1
    pattern = r"(?:'[^']+'!|\w+!)?\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?"
    matches = re.findall(pattern, formula.upper())
    return matches


def _expand_range(range_str: str) -> List[str]:
    """
    Expand a range like 'A1:A10' into individual cell addresses.
    Returns list of cell addresses.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    
    if ':' not in range_str:
        return [range_str]
    
    # Remove sheet reference if present
    if '!' in range_str:
        range_str = range_str.split('!')[-1]
    
    # Remove $ signs
    range_str = range_str.replace('$', '')
    
    try:
        start, end = range_str.split(':')
        start_col, start_row = coordinate_from_string(start)
        end_col, end_row = coordinate_from_string(end)
        
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        cells = []
        for row in range(start_row, end_row + 1):
            for col in range(start_col_idx, end_col_idx + 1):
                cells.append(f"{get_column_letter(col)}{row}")
        return cells
    except Exception:
        return [range_str]


def _normalize_sheet_name(sheet_name: str) -> str:
    normalized = (sheet_name or "").strip()
    if normalized.startswith("'") and normalized.endswith("'") and len(normalized) >= 2:
        normalized = normalized[1:-1].replace("''", "'")
    return normalized


def _split_sheet_and_range(reference: str, default_sheet: str) -> tuple[str, str]:
    if "!" not in reference:
        return default_sheet, reference

    sheet_part, range_part = reference.split("!", 1)
    return _normalize_sheet_name(sheet_part), range_part


def _collect_preflight_warnings(
    wb, sheet_name: str, changes: List[CellChange]
) -> List[CellWarning]:
    ws = wb[sheet_name]
    warnings: List[CellWarning] = []
    target_sheet_normalized = _normalize_sheet_name(sheet_name).lower()

    changing_cells: dict[str, CellChange] = {}
    for change in changes:
        if change.row < 1 or change.col < 1:
            continue
        addr = f"{get_column_letter(change.col)}{change.row}"
        changing_cells[addr] = change

    # Check 1: direct formula overwrites.
    for addr, change in changing_cells.items():
        cell = ws.cell(row=change.row, column=change.col)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            warnings.append(
                CellWarning(
                    cell=CellReference(
                        sheet=sheet_name,
                        row=change.row,
                        col=change.col,
                        address=addr,
                    ),
                    warning_type="contains_formula",
                    message=f"Cell {addr} contains a formula that will be overwritten",
                    details={
                        "current_formula": cell.value,
                        "overwrite_allowed": bool(change.forceOverwriteFormula),
                    },
                )
            )

    # Check 2: references from formulas in any worksheet.
    referenced_by: dict[str, List[str]] = {addr: [] for addr in changing_cells}
    for formula_ws in wb.worksheets:
        formula_sheet = formula_ws.title
        formula_sheet_normalized = _normalize_sheet_name(formula_sheet).lower()

        for row in formula_ws.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue

                refs = _parse_cell_references_from_formula(cell.value)
                for ref in refs:
                    ref_sheet, ref_range = _split_sheet_and_range(ref, formula_sheet)
                    if _normalize_sheet_name(ref_sheet).lower() != target_sheet_normalized:
                        continue

                    for expanded_addr in _expand_range(ref_range):
                        clean_addr = expanded_addr.replace("$", "")
                        if clean_addr in referenced_by:
                            source_addr = f"{get_column_letter(cell.column)}{cell.row}"
                            if formula_sheet_normalized != target_sheet_normalized:
                                source_addr = f"{formula_sheet}!{source_addr}"
                            referenced_by[clean_addr].append(source_addr)

    for addr, referencing_cells in referenced_by.items():
        if not referencing_cells:
            continue

        change = changing_cells[addr]
        unique_refs = sorted(set(referencing_cells))
        warnings.append(
            CellWarning(
                cell=CellReference(
                    sheet=sheet_name,
                    row=change.row,
                    col=change.col,
                    address=addr,
                ),
                warning_type="referenced_by_formula",
                message=f"Cell {addr} is referenced by {len(unique_refs)} formula(s)",
                details={
                    "referenced_by": unique_refs[:10],
                    "total_references": len(unique_refs),
                },
            )
        )

    # Check 3: workbook defined names.
    try:
        defined_names = []
        if hasattr(wb.defined_names, "values"):
            defined_names = list(wb.defined_names.values())
        elif hasattr(wb.defined_names, "definedName"):
            defined_names = list(wb.defined_names.definedName)

        for defined_name in defined_names:
            with suppress(Exception):
                for destination_sheet, destination_range in defined_name.destinations:
                    if (
                        _normalize_sheet_name(destination_sheet).lower()
                        != target_sheet_normalized
                    ):
                        continue

                    for expanded_addr in _expand_range(destination_range):
                        clean_addr = expanded_addr.replace("$", "")
                        if clean_addr not in changing_cells:
                            continue

                        change = changing_cells[clean_addr]
                        warnings.append(
                            CellWarning(
                                cell=CellReference(
                                    sheet=sheet_name,
                                    row=change.row,
                                    col=change.col,
                                    address=clean_addr,
                                ),
                                warning_type="in_named_range",
                                message=(
                                    f"Cell {clean_addr} is part of named range "
                                    f"'{defined_name.name}'"
                                ),
                                details={"named_range": defined_name.name},
                            )
                        )
    except Exception as e:
        log.warning(f"Could not check named ranges: {e}")

    return warnings


def _build_preflight_report(
    warnings: List[CellWarning],
    strict_formula_mode: bool = True,
    block_referenced_by_formula: bool = True,
    allow_formula_overwrite: bool = False,
) -> ExcelValidationResponse:
    blocked_by_formula_overwrite = False
    blocked_by_formula_reference = False

    for warning in warnings:
        details = warning.details or {}
        overwrite_allowed = bool(details.get("overwrite_allowed")) or bool(
            allow_formula_overwrite
        )

        if warning.warning_type == "contains_formula":
            if strict_formula_mode and not overwrite_allowed:
                blocked_by_formula_overwrite = True
        elif warning.warning_type == "referenced_by_formula":
            if block_referenced_by_formula:
                blocked_by_formula_reference = True

    safe_to_apply = not (blocked_by_formula_overwrite or blocked_by_formula_reference)
    if safe_to_apply:
        message = (
            f"Found {len(warnings)} warning(s)"
            if warnings
            else "No warnings - safe to apply"
        )
    else:
        reasons = []
        if blocked_by_formula_overwrite:
            reasons.append("attempted overwrite of existing formula cells")
        if blocked_by_formula_reference:
            reasons.append("changes impact cells referenced by formulas")
        message = f"Blocked by preflight: {', '.join(reasons)}"

    return ExcelValidationResponse(
        status="ok" if safe_to_apply else "blocked",
        warnings=warnings,
        safe_to_apply=safe_to_apply,
        message=message,
    )


@router.post("/validate")
async def validate_excel_changes(
    request: ExcelValidationRequest,
    user=Depends(get_verified_user),
) -> ExcelValidationResponse:
    """
    Validate proposed changes before applying them.
    
    Returns warnings about:
    - Cells that contain formulas (will be overwritten)
    - Cells that are referenced by other formulas (may break calculations)
    - Cells within named ranges
    
    This endpoint is advisory - changes can still be applied even with warnings.
    """
    try:
        # Get the file from database
        file = Files.get_file_by_id(request.fileId)

        if not file:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=ERROR_MESSAGES.NOT_FOUND,
            )

        # Check user permissions
        if file.user_id != user.id and user.role != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=ERROR_MESSAGES.ACCESS_PROHIBITED,
            )

        # Get the file path from storage
        file_path = Storage.get_file(file.path)
        file_path = Path(file_path)

        if not file_path.is_file():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found in storage",
            )

        # Load the workbook (need formulas, not values)
        try:
            wb = openpyxl.load_workbook(
                file_path, **_get_workbook_load_kwargs(file.filename)
            )
        except Exception as e:
            log.error(f"Error loading workbook for validation {request.fileId}: {e}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid Excel file: {str(e)}",
            )

        # Check if sheet exists
        if request.sheet not in wb.sheetnames:
            wb.close()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Sheet '{request.sheet}' not found. Available: {', '.join(wb.sheetnames)}",
            )

        try:
            warnings = _collect_preflight_warnings(
                wb=wb,
                sheet_name=request.sheet,
                changes=request.changes,
            )
            return _build_preflight_report(
                warnings=warnings,
                strict_formula_mode=bool(request.strictFormulaMode),
                block_referenced_by_formula=bool(request.blockReferencedByFormula),
                allow_formula_overwrite=bool(request.allowFormulaOverwrite),
            )
        finally:
            wb.close()

    except HTTPException:
        raise
    except Exception as e:
        log.error(f"Unexpected error validating Excel changes: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Internal server error: {str(e)}",
        )
