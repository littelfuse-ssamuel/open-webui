from open_webui.routers.images import (
    get_image_data,
    upload_image,
)

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Request,
    UploadFile,
)
from typing import Optional
from pathlib import Path

from open_webui.storage.provider import Storage

from open_webui.models.chats import Chats
from open_webui.models.files import Files
from open_webui.routers.files import upload_file_handler

import mimetypes
import base64
import io
import os
import re
import logging
import requests

# Littelfuse custom import for Excel handling
import openpyxl

log = logging.getLogger(__name__)

BASE64_IMAGE_URL_PREFIX = re.compile(r"data:image/\w+;base64,", re.IGNORECASE)
MARKDOWN_IMAGE_URL_PATTERN = re.compile(r"!\[(.*?)\]\((.+?)\)", re.IGNORECASE)


def get_image_base64_from_url(url: str) -> Optional[str]:
    """
    Convert an image URL or file ID to a base64 data URI.
    
    Args:
        url: Either an HTTP(S) URL or a file ID
        
    Returns:
        Base64 data URI string or None if failed
    """
    try:
        if url.startswith("http"):
            # Download the image from the URL
            response = requests.get(url)
            response.raise_for_status()
            image_data = response.content
            encoded_string = base64.b64encode(image_data).decode("utf-8")
            content_type = response.headers.get("Content-Type", "image/png")
            return f"data:{content_type};base64,{encoded_string}"
        else:
            file = Files.get_file_by_id(url)

            if not file:
                return None

            file_path = Storage.get_file(file.path)
            file_path = Path(file_path)

            if file_path.is_file():
                with open(file_path, "rb") as image_file:
                    encoded_string = base64.b64encode(image_file.read()).decode("utf-8")
                    content_type, _ = mimetypes.guess_type(file_path.name)
                    return f"data:{content_type};base64,{encoded_string}"
            else:
                return None

    except Exception as e:
        log.error(f"Error getting image base64 from URL {url}: {e}")
        return None


def get_image_url_from_base64(request, base64_image_string, metadata, user):
    """
    Upload a base64 image and return its URL.
    
    Args:
        request: FastAPI request object
        base64_image_string: Base64 encoded image with data URI prefix
        metadata: Metadata dict for the file
        user: User object
        
    Returns:
        URL string or None if failed
    """
    if BASE64_IMAGE_URL_PREFIX.match(base64_image_string):
        image_url = ""
        # Extract base64 image data from the line
        image_data, content_type = get_image_data(base64_image_string)
        if image_data is not None:
            _, image_url = upload_image(
                request,
                image_data,
                content_type,
                metadata,
                user,
            )

        return image_url
    return None


def convert_markdown_base64_images(request, content: str, metadata, user):
    """
    Convert base64 images in markdown to uploaded file URLs.
    
    Args:
        request: FastAPI request object
        content: Markdown content with potential base64 images
        metadata: Metadata dict for uploaded files
        user: User object
        
    Returns:
        Markdown content with base64 images replaced by URLs
    """
    def replace(match):
        base64_string = match.group(2)
        MIN_REPLACEMENT_URL_LENGTH = 1024
        if len(base64_string) > MIN_REPLACEMENT_URL_LENGTH:
            url = get_image_url_from_base64(request, base64_string, metadata, user)
            if url:
                return f"![{match.group(1)}]({url})"
        return match.group(0)

    return MARKDOWN_IMAGE_URL_PATTERN.sub(replace, content)


def load_b64_audio_data(b64_str):
    """
    Decode base64 audio data.
    
    Args:
        b64_str: Base64 encoded audio with optional data URI prefix
        
    Returns:
        Tuple of (audio_data bytes, content_type string) or (None, None) on error
    """
    try:
        if "," in b64_str:
            header, b64_data = b64_str.split(",", 1)
        else:
            b64_data = b64_str
            header = "data:audio/wav;base64"
        audio_data = base64.b64decode(b64_data)
        content_type = (
            header.split(";")[0].split(":")[1] if ";" in header else "audio/wav"
        )
        return audio_data, content_type
    except Exception as e:
        log.error(f"Error decoding base64 audio data: {e}")
        return None, None


def upload_audio(request, audio_data, content_type, metadata, user):
    """
    Upload audio data and return its URL.
    
    Args:
        request: FastAPI request object
        audio_data: Raw audio bytes
        content_type: MIME type of the audio
        metadata: Metadata dict for the file
        user: User object
        
    Returns:
        URL string for the uploaded audio
    """
    audio_format = mimetypes.guess_extension(content_type)
    file = UploadFile(
        file=io.BytesIO(audio_data),
        filename=f"generated-{audio_format}",  # will be converted to a unique ID on upload_file
        headers={
            "content-type": content_type,
        },
    )
    file_item = upload_file_handler(
        request,
        file=file,
        metadata=metadata,
        process=False,
        user=user,
    )
    url = request.app.url_path_for("get_file_content_by_id", id=file_item.id)
    return url


def get_audio_url_from_base64(request, base64_audio_string, metadata, user):
    """
    Upload a base64 audio file and return its URL.
    
    Args:
        request: FastAPI request object
        base64_audio_string: Base64 encoded audio with data URI prefix
        metadata: Metadata dict for the file
        user: User object
        
    Returns:
        URL string or None if failed
    """
    if "data:audio/wav;base64" in base64_audio_string:
        audio_url = ""
        # Extract base64 audio data from the line
        audio_data, content_type = load_b64_audio_data(base64_audio_string)
        if audio_data is not None:
            audio_url = upload_audio(
                request,
                audio_data,
                content_type,
                metadata,
                user,
            )
        return audio_url
    return None


def get_file_url_from_base64(request, base64_file_string, metadata, user):
    """
    Upload a base64 file (image or audio) and return its URL.
    
    Args:
        request: FastAPI request object
        base64_file_string: Base64 encoded file with data URI prefix
        metadata: Metadata dict for the file
        user: User object
        
    Returns:
        URL string or None if not a supported type
    """
    if BASE64_IMAGE_URL_PREFIX.match(base64_file_string):
        return get_image_url_from_base64(request, base64_file_string, metadata, user)
    elif "data:audio/wav;base64" in base64_file_string:
        return get_audio_url_from_base64(request, base64_file_string, metadata, user)
    return None


def get_image_base64_from_file_id(id: str) -> Optional[str]:
    """
    Get base64 data URI for an image file by its ID.
    
    Args:
        id: File ID
        
    Returns:
        Base64 data URI string or None if not found
    """
    file = Files.get_file_by_id(id)
    if not file:
        return None

    try:
        file_path = Storage.get_file(file.path)
        file_path = Path(file_path)

        # Check if the file exists
        if file_path.is_file():
            with open(file_path, "rb") as image_file:
                encoded_string = base64.b64encode(image_file.read()).decode("utf-8")
                content_type, _ = mimetypes.guess_type(file_path.name)
                return f"data:{content_type};base64,{encoded_string}"
        else:
            return None
    except Exception as e:
        log.error(f"Error getting image base64 from file ID {id}: {e}")
        return None


# ===========================================
# Custom methods (Littelfuse additions)
# ===========================================


def upload_excel_file(request, file_path, metadata, user):
    """
    Upload an Excel file from a file path and return the file artifact dict.
    
    Littelfuse custom method.

    Args:
        request: FastAPI request object
        file_path: Path to the Excel file
        metadata: Metadata dict for the file
        user: User object

    Returns:
        Dict with file artifact structure or None if failed
    """
    try:
        if not os.path.exists(file_path):
            log.error(f"Excel file not found: {file_path}")
            return None

        # Read the file
        with open(file_path, "rb") as f:
            excel_data = f.read()

        # Extract sheet names using openpyxl
        sheet_names = []
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            log.warning(f"Could not extract sheet names from {file_path}: {e}")

        # Create UploadFile object
        filename = os.path.basename(file_path)
        file = UploadFile(
            file=io.BytesIO(excel_data),
            filename=filename,
            headers={
                "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
        )

        # Add sheet names to metadata if available
        if metadata is None:
            metadata = {}
        if sheet_names:
            metadata["sheetNames"] = sheet_names

        # Upload the file
        file_item = upload_file_handler(
            request,
            file=file,
            metadata=metadata,
            process=False,
            user=user,
        )

        # Build file artifact dict
        url = request.app.url_path_for("get_file_content_by_id", id=file_item.id)

        artifact = {
            "type": "excel",
            "url": url,
            "name": filename,
            "fileId": str(file_item.id),
            "meta": {
                "sheetNames": sheet_names,
                "activeSheet": sheet_names[0] if sheet_names else None,
            },
        }

        log.info(f"Uploaded Excel file: {filename} -> {file_item.id}")
        return artifact

    except Exception as e:
        log.error(f"Error uploading Excel file {file_path}: {e}")
        return None


def get_excel_artifact_from_base64(request, base64_excel_string, filename, metadata, user):
    """
    Convert base64 Excel data to file artifact dict.
    
    Littelfuse custom method.

    Args:
        request: FastAPI request object
        base64_excel_string: Base64 encoded Excel data with data URI prefix
        filename: Filename for the Excel file
        metadata: Metadata dict for the file
        user: User object

    Returns:
        Dict with file artifact structure or None if failed
    """
    log.info(
        f"get_excel_artifact_from_base64 called with filename={filename}, "
        f"string_length={len(base64_excel_string)}"
    )

    if "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64," in base64_excel_string:
        try:
            # Extract base64 data
            excel_data_base64 = base64_excel_string.split(",", 1)[1]
            log.info(f"Extracted base64 data, length={len(excel_data_base64)}")

            excel_data = base64.b64decode(excel_data_base64)
            log.info(f"Decoded Excel data, byte length={len(excel_data)}")

            # Extract sheet names using openpyxl
            sheet_names = []
            try:
                wb = openpyxl.load_workbook(io.BytesIO(excel_data), read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
            except Exception as e:
                log.warning(f"Could not extract sheet names from base64 Excel data: {e}")

            # Create UploadFile object
            file = UploadFile(
                file=io.BytesIO(excel_data),
                filename=filename if filename else "output.xlsx",
                headers={
                    "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                },
            )

            # Add sheet names to metadata
            if metadata is None:
                metadata = {}
            if sheet_names:
                metadata["sheetNames"] = sheet_names

            # Upload the file
            log.info("Uploading Excel file to storage...")
            file_item = upload_file_handler(
                request,
                file=file,
                metadata=metadata,
                process=False,
                user=user,
            )
            log.info(f"File uploaded successfully, file_id={file_item.id}")

            # Build file artifact dict
            url = request.app.url_path_for("get_file_content_by_id", id=file_item.id)

            artifact = {
                "type": "excel",
                "url": url,
                "name": filename if filename else "output.xlsx",
                "fileId": str(file_item.id),
                "meta": {
                    "sheetNames": sheet_names,
                    "activeSheet": sheet_names[0] if sheet_names else None,
                },
            }

            log.info(
                f"Successfully created Excel artifact from base64: "
                f"{filename if filename else 'output.xlsx'} -> {file_item.id}, url={url}"
            )
            return artifact

        except Exception as e:
            log.error(f"Error processing base64 Excel data: {e}", exc_info=True)
            return None

    log.warning("Base64 string does not contain Excel data URI prefix")
    return None